import re
import json
import base64
import hashlib
from pathlib import Path
from io import BytesIO
from zipfile import ZipFile
from datetime import datetime

import streamlit as st
import pandas as pd
import requests
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side, Protection
from openpyxl.worksheet.datavalidation import DataValidation
from pathlib import Path
from datetime import datetime
from pathlib import Path

def get_version():
    try:
        return Path("version.txt").read_text().strip()
    except:
        return "?"

version = get_version()
today = datetime.now().strftime("%Y-%m-%d")

def get_version():
    try:
        return Path("version.txt").read_text().strip()
    except:
        return "?"

# =========================
# إعدادات عامة
# =========================
st.set_page_config(page_title="مركز التصحيح المركزي", layout="wide")

st.markdown("""
<style>

/* 🔥 تكبير عنوان السؤال */
.question-title {
    font-size: 26px !important;
    font-weight: bold;
    color: #1f3b73;
}

/* ✨ نص السؤال */
.question-text {
    font-size: 22px !important;
    color: #444;
}

/* 💚 عمود Points */
.points-text {
    font-size: 24px !important;
    color: #2e7d32;
    font-weight: bold;
}

/* 📦 الكرت كامل */
.question-box {
    padding: 20px;
    border-radius: 12px;
    background-color: #f8f9fb;
    margin-bottom: 20px;
}

</style>
""", unsafe_allow_html=True)

APP_DATA_DIR = Path.home() / "Jidhafs_Control_Center_Data"
APP_DATA_DIR.mkdir(exist_ok=True)

LOCAL_REPORT_FILE = APP_DATA_DIR / "reports.xlsx"
GRADE_TEMPLATES_FILE = APP_DATA_DIR / "grade_templates.json"

# ضعي رابط Google Apps Script هنا لاحقًا
GOOGLE_SCRIPT_URL = "https://script.google.com/macros/s/AKfycby1miY6jZcCNqtYtWZzaSkDVZRSOyeKzl2eN9aGrlnLkRCEC619vL8eLzYstP9pieKo0w/exec"

# كلمة مرور الأدمن
ADMIN_PASSWORD = "Jidhafs!1825"

# كلمة مرور موحدة للمستخدمين العاديين
USER_PASSWORD = "User!1234"


# =========================
# أدوات مساعدة
# =========================
def clean_header(header):
    header = str(header or "")
    header = header.replace("\xa0", " ")
    header = " ".join(header.split())
    return header


def safe_filename(name):
    name = clean_header(name)
    name = re.sub(r'[\\/:*?"<>|]', "-", name)
    return name.strip() or "ملف"


SCHOOL_OPTIONS = [
    "مدرسة النور الثانوية للبنات",
    "مدرسة المعرفة الثانوية للبنات",
    "مدرسة الرفاع الغربي الثانوية للبنات",
    "مدرسة جدحفص الثانوية للبنات",
]


def detect_school_column(df):
    """يرجع اسم عمود المدرسة فقط إذا كان الهيدر مطابقًا لاسم عمود مدرسة واضح.
    مهم: لا نستخدم البحث الجزئي حتى لا يعتبر سؤالًا فيه كلمة school أنه عمود مدرسة.
    """
    school_headers = {
        "school",
        "school name",
        "student school",
        "student's school",
        "student’s school",
        "اسم المدرسة",
        "المدرسة",
        "مدرسة الطالبة",
        "المدرسة التابع لها",
        "المدرسة التابعة لها",
        "المدرسة الأصلية",
        "المدرسة الاصلية",
    }

    for col in df.columns:
        h = clean_header(col).lower()
        if h in school_headers:
            return col
    return None

def arabic_day_name(dt):
    days = {
        "Saturday": "السبت",
        "Sunday": "الأحد",
        "Monday": "الإثنين",
        "Tuesday": "الثلاثاء",
        "Wednesday": "الأربعاء",
        "Thursday": "الخميس",
        "Friday": "الجمعة",
    }
    return days.get(dt.strftime("%A"), dt.strftime("%A"))


def get_logo_base64():
    logo_path = Path("logo.png")
    if logo_path.exists():
        with open(logo_path, "rb") as f:
            return base64.b64encode(f.read()).decode()
    return None


def load_grade_templates():
    if GRADE_TEMPLATES_FILE.exists():
        try:
            with open(GRADE_TEMPLATES_FILE, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            return {}
    return {}


def save_grade_templates(templates):
    with open(GRADE_TEMPLATES_FILE, "w", encoding="utf-8") as f:
        json.dump(templates, f, ensure_ascii=False, indent=2)


def get_file_signature(df):
    cols = "||".join([clean_header(c) for c in df.columns])
    return hashlib.md5(cols.encode("utf-8")).hexdigest()


# =========================
# تسجيل الدخول
# =========================
def login_screen():
    if "logged_in" not in st.session_state:
        st.session_state.logged_in = False
        st.session_state.role = None
        st.session_state.username = None

    if st.session_state.logged_in:
        return True

    st.markdown(f"""
    <link href="https://fonts.googleapis.com/css2?family=Cairo:wght@300;400;600;700;900&family=Tajawal:wght@300;400;500;700;800&display=swap" rel="stylesheet">
    <style>
    :root {{
        --primary: #0d3060;
        --primary-light: #1a4f8a;
        --accent: #c8a035;
        --border: #d6e4f7;
        --text-main: #1a2a45;
        --text-muted: #5a7092;
        --radius: 16px;
        --radius-sm: 10px;
    }}
    html, body, [class*="css"] {{
        direction: rtl !important;
        text-align: right !important;
        font-family: 'Cairo', 'Tajawal', sans-serif !important;
        background: linear-gradient(135deg, #e8f0fb 0%, #f4f7fc 60%, #eef3fb 100%) !important;
    }}
    .block-container {{
        max-width: 480px !important;
        margin: auto !important;
        padding-top: 3rem !important;
    }}
    .login-card {{
        background: #ffffff;
        border-radius: var(--radius);
        padding: 40px 36px 28px;
        box-shadow: 0 8px 32px rgba(13,48,96,0.14);
        border: 1px solid var(--border);
        margin-bottom: 20px;
        text-align: center;
    }}
    .login-logo-ring {{
        width: 78px; height: 78px;
        background: linear-gradient(135deg, var(--primary) 0%, var(--primary-light) 100%);
        border-radius: 50%;
        display: flex; align-items: center; justify-content: center;
        margin: 0 auto 18px auto;
        font-size: 32px;
        box-shadow: 0 6px 20px rgba(13,48,96,0.25);
    }}
    .login-title {{
        font-family: 'Cairo', sans-serif;
        font-size: 22px;
        font-weight: 900;
        color: var(--primary);
        margin: 0 0 4px 0;
    }}
    .login-subtitle {{
        font-size: 13px;
        color: var(--text-muted);
        font-weight: 500;
        margin-bottom: 12px;
    }}
    .login-date-badge {{
        display: inline-block;
        background: linear-gradient(135deg, #eaf3ff, #f4f9ff);
        border: 1px solid var(--border);
        color: var(--primary);
        font-size: 12px;
        font-weight: 700;
        padding: 4px 14px;
        border-radius: 20px;
        margin-bottom: 4px;
    }}
    label {{
        font-family: 'Cairo', sans-serif !important;
        font-size: 14px !important;
        font-weight: 700 !important;
        color: var(--text-main) !important;
    }}
    .stTextInput > div > div > input {{
        font-family: 'Cairo', sans-serif !important;
        font-size: 15px !important;
        border: 1.5px solid var(--border) !important;
        border-radius: var(--radius-sm) !important;
        padding: 10px 14px !important;
        direction: rtl !important;
    }}
    .stButton > button {{
        font-family: 'Cairo', sans-serif !important;
        font-size: 16px !important;
        font-weight: 800 !important;
        background: linear-gradient(135deg, var(--primary) 0%, var(--primary-light) 100%) !important;
        color: #fff !important;
        border: none !important;
        border-radius: var(--radius-sm) !important;
        padding: 12px 28px !important;
        width: 100% !important;
        box-shadow: 0 4px 16px rgba(13,48,96,0.28) !important;
        transition: all 0.22s !important;
    }}
    .stButton > button:hover {{
        transform: translateY(-2px) !important;
        box-shadow: 0 8px 24px rgba(13,48,96,0.36) !important;
    }}
    .stRadio > div {{
        direction: rtl !important;
        background: #f4f8ff !important;
        border-radius: var(--radius-sm) !important;
        padding: 8px !important;
        border: 1px solid var(--border) !important;
    }}
    [data-testid="stAlert"] {{
        border-radius: var(--radius-sm) !important;
        font-family: 'Cairo', sans-serif !important;
        font-size: 14px !important;
        font-weight: 600 !important;
    }}
    </style>
    <div class="login-card">
        <div class="login-logo-ring">🏫</div>
        <div class="login-title">مركز التصحيح المركزي</div>
        <div class="login-subtitle">مدرسة جدحفص الثانوية للبنات</div>
        <div class="login-date-badge">📅 {today}</div>
    </div>
    """, unsafe_allow_html=True)

    role = st.radio("نوع الدخول", ["مستخدم", "أدمن"], horizontal=True)

    if role == "مستخدم":
        username = st.text_input("اسم المستخدم / اسم المعلم")
        password = st.text_input("كلمة المرور", type="password")

        if st.button("دخول"):
            if username.strip() and password == USER_PASSWORD:
                st.session_state.logged_in = True
                st.session_state.role = "user"
                st.session_state.username = username.strip()
                st.rerun()
            else:
                st.error("الاسم أو كلمة المرور غير صحيحة ❌")

    else:
        username = st.text_input("اسم الأدمن", value="Admin")
        password = st.text_input("كلمة مرور الأدمن", type="password")

        if st.button("دخول الأدمن"):
            if password == ADMIN_PASSWORD:
                st.session_state.logged_in = True
                st.session_state.role = "admin"
                st.session_state.username = username.strip() or "Admin"
                st.rerun()
            else:
                st.error("كلمة مرور الأدمن غير صحيحة ❌")

    return False


# =========================
# التقرير المحلي + Google Sheet
# =========================
def append_local_report(record):
    columns = [
        "اليوم",
        "التاريخ",
        "الساعة",
        "اسم المستخدم",
        "نوع المستخدم",
        "نوع العملية",
        "اسم الملف الأصلي",
        "اسم الملف الجديد",
        "أسماء الأجزاء",
        "عدد الأجزاء",
        "عدد الاستجابات",
        "عدد الاستجابات في كل جزء",
        "هل تم التقسيم",
        "هل تم التجميع",
        "اسم ملف التجميع",
        "الحالة",
        "ملاحظات",
    ]

    df_new = pd.DataFrame([record], columns=columns)

    if LOCAL_REPORT_FILE.exists():
        try:
            df_old = pd.read_excel(LOCAL_REPORT_FILE)
            df_all = pd.concat([df_old, df_new], ignore_index=True)
        except Exception:
            df_all = df_new
    else:
        df_all = df_new

    df_all.to_excel(LOCAL_REPORT_FILE, index=False)


def send_report_to_google(record):
    if not GOOGLE_SCRIPT_URL or GOOGLE_SCRIPT_URL == "PUT_YOUR_GOOGLE_SCRIPT_URL_HERE":
        return False

    try:
        response = requests.post(GOOGLE_SCRIPT_URL, json=record, timeout=8)
        return response.status_code in [200, 201]
    except Exception:
        return False


def log_operation(
    operation_type,
    original_file="",
    new_file_name="",
    part_names=None,
    part_count=0,
    response_count=0,
    responses_per_part="",
    split_done="لا",
    merge_done="لا",
    merged_file_name="",
    status="تم",
    notes="",
):
    now = datetime.now()
    record = {
        "اليوم": arabic_day_name(now),
        "التاريخ": now.strftime("%Y-%m-%d"),
        "الساعة": now.strftime("%H:%M:%S"),
        "اسم المستخدم": st.session_state.get("username", ""),
        "نوع المستخدم": st.session_state.get("role", ""),
        "نوع العملية": operation_type,
        "اسم الملف الأصلي": original_file,
        "اسم الملف الجديد": new_file_name,
        "أسماء الأجزاء": "، ".join(part_names or []),
        "عدد الأجزاء": part_count,
        "عدد الاستجابات": response_count,
        "عدد الاستجابات في كل جزء": responses_per_part,
        "هل تم التقسيم": split_done,
        "هل تم التجميع": merge_done,
        "اسم ملف التجميع": merged_file_name,
        "الحالة": status,
        "ملاحظات": notes,
    }

    append_local_report(record)
    send_report_to_google(record)


# =========================
# إعدادات الأعمدة والدرجات
# =========================
hidden_headers = [
    "id",
    "start time",
    "completion time",
    "email",
    "name",
    "grade posted time",
    "last modified time",
    "school",
    "المدرسة",
    "اسم المدرسة",
    "الاسم الرباعي",
    "الرقم الأكاديمي",
    "الشعبة",
    "points - الاسم الرباعي",
    "points - الرقم الأكاديمي",
    "points - الشعبة",
]


def _is_student_info_field(header):
    """هل هذا الحقل متعلق ببيانات الطالبة/الطالب (اسم، رقم، شعبة) وليس سؤالاً حقيقياً؟"""
    h = clean_header(header).lower()
    return any(x in h for x in [
        # اسم — جميع الصياغات
        "اسم الطالبة", "اسم الطالب", "اسم الطلاب", "اسم الطالبات",
        "الاسم الرباعي", "الاسم الثلاثي", "الاسم الكامل",
        "بيانات الطالبة", "بيانات الطالب", "بيانات الطلاب",
        "الاسم والرقم", "اسم ورقم",
        # رقم — جميع الصياغات
        "الرقم الأكاديمي", "رقم الاكاديمي", "رقم أكاديمي",
        "رقم الطالبة", "رقم الطالب", "الرقم المدرسي",
        # شعبة / فصل
        "الشعبة", "شعبة", "الفصل الدراسي", "الفصل",
        # إنجليزي
        "student name", "student id", "student number",
        "class section", "section", "grade level",
    ])


def should_auto_hide(header):
    h = clean_header(header).lower()
    hidden = [
        "id", "start time", "completion time", "email", "name",
        "grade posted time", "last modified time",
        "school", "school name", "student school", "student's school", "student’s school",
        "المدرسة", "اسم المدرسة", "مدرسة الطالبة", "المدرسة التابع لها",
        "المدرسة التابعة لها", "المدرسة الأصلية", "المدرسة الاصلية",
    ]
    return (
        "feedback" in h
        or h in hidden
        or _is_student_info_field(header)
    )


def is_points_column(header):
    header_lower = clean_header(header).lower()
    if _is_student_info_field(header):
        return False
    return (
        "points" in header_lower
        and header_lower != "total points"
    )


def find_related_question(points_header, all_headers):
    points_header_clean = clean_header(points_header)

    possible_question = re.sub(r"^points[ ]*-[ ]*", "", points_header_clean, flags=re.IGNORECASE).strip()
    if possible_question and possible_question != points_header_clean:
        for h in all_headers:
            if clean_header(h) == possible_question:
                return clean_header(h)
        return possible_question

    return points_header_clean


def extract_score_from_text(text):
    """يستخرج مجموع الدرجات الظاهرة من نص السؤال (مثل: 5 درجات، ٥ درجات)
    ويجمعها إذا كان في أكثر من بند."""
    text = str(text or "")

    # تحويل الأرقام العربية إلى إنجليزية
    arabic_digits = str.maketrans("٠١٢٣٤٥٦٧٨٩", "0123456789")
    text = text.translate(arabic_digits)

    # استخراج كل الأرقام المرتبطة بكلمة درجة/درجات
    matches = re.findall(r"([0-9]+(?:[.][0-9]+)?)[ ]*درجات?", text)

    if matches:
        total = sum(float(x) for x in matches)
        return int(total) if total.is_integer() else total

    return None


def detect_max_scores_from_data(df):
    max_scores = {}
    unknown_points = []
    all_headers = list(df.columns)

    for col in df.columns:
        header = clean_header(col)
        if not is_points_column(header):
            continue

        values = pd.to_numeric(df[col], errors="coerce").dropna()
        if len(values) > 0:
            max_value = float(values.max())
            if max_value > 0:
                if max_value.is_integer():
                    max_value = int(max_value)
                max_scores[header] = max_value
            else:
                unknown_points.append({
                    "عمود الدرجة": header,
                    "السؤال": find_related_question(header, all_headers),
                    "الدرجة الكبرى": None,
                })
        else:
            unknown_points.append({
                "عمود الدرجة": header,
                "السؤال": find_related_question(header, all_headers),
                "الدرجة الكبرى": None,
            })

    return max_scores, unknown_points


def update_points_headers_with_max(ws, max_scores):
    for col in ws.columns:
        cell = col[0]
        header = clean_header(cell.value)
        if header in max_scores:
            if "الدرجة من" not in header:
                cell.value = f"{header} / الدرجة من {max_scores[header]}"


def add_score_validation(ws, max_scores):
    for col in ws.columns:
        col_letter = col[0].column_letter
        header = clean_header(col[0].value)
        original_header = header.split(" / الدرجة من ")[0].strip()

        if original_header in max_scores:
            max_score = max_scores[original_header]
            dv = DataValidation(
                type="decimal",
                operator="between",
                formula1="0",
                formula2=str(max_score),
                allow_blank=True,
            )
            dv.errorTitle = "خطأ في الدرجة"
            dv.error = f"الدرجة المدخلة أكبر من الدرجة المسموح بها. الحد الأعلى هو {max_score}."
            dv.promptTitle = "تنبيه"
            dv.prompt = f"اكتب درجة من 0 إلى {max_score} فقط."
            dv.showErrorMessage = True
            dv.showInputMessage = True
            ws.add_data_validation(dv)
            dv.add(f"{col_letter}2:{col_letter}{ws.max_row}")


# =========================
# تنسيق Excel
# =========================
def format_excel_file(
    excel_bytes,
    lock_sheet=False,
    merge_mode=False,
    extra_hidden_columns=None,
    max_scores=None,
    excluded_columns=None,
    corrector_name="",
    auditor_name="",
    school_name="مدرسة جدحفص الثانوية للبنات",
):
    if extra_hidden_columns is None:
        extra_hidden_columns = []
    if max_scores is None:
        max_scores = {}
    if excluded_columns is None:
        excluded_columns = []

    extra_hidden_columns = [clean_header(c) for c in extra_hidden_columns]
    excluded_columns = [clean_header(c) for c in excluded_columns]

    excel_bytes.seek(0)
    wb = load_workbook(excel_bytes)
    ws = wb.active

    header_fill = PatternFill("solid", fgColor="BFEFFF")
    points_fill = PatternFill("solid", fgColor="CCFFCC")
    total_fill = PatternFill("solid", fgColor="FFF2CC")

    thin_border = Border(
        left=Side(style="thin", color="808080"),
        right=Side(style="thin", color="808080"),
        top=Side(style="thin", color="808080"),
        bottom=Side(style="thin", color="808080"),
    )

    visible_widths = {}
    points_cols_for_total = []
    total_col_letter = None

    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
            cell.font = Font(size=18)
            cell.border = thin_border
            if lock_sheet:
                cell.protection = Protection(locked=True)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = Font(size=22, bold=True)

    update_points_headers_with_max(ws, max_scores)

    for col in ws.columns:
        col_letter = col[0].column_letter
        header = clean_header(col[0].value)
        original_header = header.split(" / الدرجة من ")[0].strip()
        header_lower = original_header.lower()

        ws.column_dimensions[col_letter].hidden = False

        if header_lower == "total points":
            total_col_letter = col_letter
            for cell in col:
                cell.fill = total_fill

        # أعمدة الدرجات — نستخدم is_points_column الذي يستثني بيانات الطالبة تلقائياً
        if is_points_column(original_header):
            if merge_mode:
                # في التجميع لا نعتمد على max_scores؛ نجمع الدرجات المكتوبة فعليًا في ملفات المصححات
                points_cols_for_total.append(col_letter)
                for cell in col:
                    cell.fill = points_fill

            elif original_header in max_scores:
                points_cols_for_total.append(col_letter)

                # الهيدر وخلايا الدرجات تكون خضراء افتراضيًا
                # والتلوين الشرطي يحول الخلية للأحمر إذا صارت فاضية
                for cell in col:
                    cell.fill = points_fill

                if lock_sheet:
                    for cell in col[1:]:
                        cell.protection = Protection(locked=False)

            elif original_header in excluded_columns:
                # مؤكد يدويًا أنه لا يحتاج درجة: نخفيه ونقفله
                ws.column_dimensions[col_letter].hidden = True
                for cell in col:
                    cell.protection = Protection(locked=True)
                continue

            else:
                # غير معتمد وغير مؤكد: نخفيه احتياطًا ولا يدخل في المجموع
                ws.column_dimensions[col_letter].hidden = True
                for cell in col:
                    cell.protection = Protection(locked=True)
                continue

        if merge_mode:
            if header_lower == "رقم":
                ws.column_dimensions[col_letter].hidden = True
        else:
            if original_header in extra_hidden_columns or should_auto_hide(original_header):
                ws.column_dimensions[col_letter].hidden = True
                continue

        max_len = 0
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))

        if col_letter == "A":
            width = 8
        elif max_len <= 20:
            width = 18
        elif max_len <= 60:
            width = 28
        elif max_len <= 140:
            width = 38
        else:
            width = 50

        ws.column_dimensions[col_letter].width = width
        visible_widths[col_letter] = width

    if total_col_letter and points_cols_for_total:
        for row_num in range(2, ws.max_row + 1):
            formula = "+".join([f"{col}{row_num}" for col in points_cols_for_total])
            ws[f"{total_col_letter}{row_num}"] = f"={formula}"
            ws[f"{total_col_letter}{row_num}"].fill = total_fill
            if lock_sheet:
                ws[f"{total_col_letter}{row_num}"].protection = Protection(locked=True)

    add_score_validation(ws, max_scores)

    # تلوين خلايا الدرجات حسب الحالة:
    # فارغة = أحمر، فيها درجة = أخضر
    # هذا الجزء يضيف Conditional Formatting فعلي داخل ملف Excel.
    if not merge_mode:
        from openpyxl.formatting.rule import FormulaRule
        from openpyxl.styles import PatternFill as PF

        red_fill_cf = PF(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
        green_fill_cf = PF(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")

        for col in ws.columns:
            col_letter = col[0].column_letter
            header = clean_header(col[0].value)
            original_header = header.split(" / الدرجة من ")[0].strip()

            if is_points_column(original_header) and not ws.column_dimensions[col_letter].hidden:
                start_row = 2
                end_row = ws.max_row
                if end_row < start_row:
                    continue

                cell_range = f"{col_letter}{start_row}:{col_letter}{end_row}"

                # لون مبدئي حسب القيم الحالية داخل الملف
                for cell in ws[f"{col_letter}{start_row}:{col_letter}{end_row}"]:
                    c = cell[0]
                    if c.value is None or str(c.value).strip() == "":
                        c.fill = red_fill_cf
                    else:
                        c.fill = green_fill_cf

                # 🔴 إذا الخلية فاضية أو فيها مسافات فقط
                # الصيغة نسبية: تبدأ من الصف 2، وExcel يطبقها على باقي الصفوف تلقائيًا.
                ws.conditional_formatting.add(
                    cell_range,
                    FormulaRule(
                        formula=[f'LEN(TRIM({col_letter}{start_row}&""))=0'],
                        fill=red_fill_cf,
                        stopIfTrue=True,
                    )
                )

                # 🟢 إذا الخلية فيها قيمة
                ws.conditional_formatting.add(
                    cell_range,
                    FormulaRule(
                        formula=[f'LEN(TRIM({col_letter}{start_row}&""))>0'],
                        fill=green_fill_cf,
                    )
                )

    # إضافة ورقة معلومات المصححة والمدققة للملفات المقسمة
    # المطلوب: تظهر خانات فقط، وتُكتب الأسماء يدويًا داخل ملف Excel بعد التنزيل.
    if not merge_mode:
        if "معلومات" not in wb.sheetnames:
            ws_info = wb.create_sheet("معلومات")
            ws_info.sheet_view.rightToLeft = True

            info_header_fill = PatternFill("solid", fgColor="15396B")
            info_fill        = PatternFill("solid", fgColor="EAF6FF")
            thin_info_border = Border(
                left=Side(style="thin", color="AAAAAA"),
                right=Side(style="thin", color="AAAAAA"),
                top=Side(style="thin", color="AAAAAA"),
                bottom=Side(style="thin", color="AAAAAA"),
            )
            bold_white  = Font(bold=True, size=13, color="FFFFFF")
            center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

            headers = ["اسم المصححة", "اسم المدققة"]
            for col_idx, header_text in enumerate(headers, start=2):
                c_header = ws_info.cell(row=2, column=col_idx, value=header_text)
                c_header.fill = info_header_fill
                c_header.font = bold_white
                c_header.alignment = center_align
                c_header.border = thin_info_border

                c_blank = ws_info.cell(row=3, column=col_idx, value="")
                c_blank.fill = info_fill
                c_blank.alignment = center_align
                c_blank.border = thin_info_border

                ws_info.column_dimensions[c_header.column_letter].width = 32

            ws_info.row_dimensions[2].height = 34
            ws_info.row_dimensions[3].height = 34
            ws_info.sheet_state = "visible"
        wb.active = wb["Responses"] if "Responses" in wb.sheetnames else ws

    for row in ws.iter_rows():
        row_num = row[0].row
        if row_num == 1:
            ws.row_dimensions[row_num].height = 60
            continue

        needed_height = 35
        for cell in row:
            col_letter = cell.column_letter
            if ws.column_dimensions[col_letter].hidden:
                continue
            if not cell.value:
                continue

            text = str(cell.value).strip()
            if len(text) < 20:
                continue

            width = visible_widths.get(col_letter, 30)
            chars_per_line = max(int(width * 1.35), 18)
            estimated_lines = text.count("\n") + (len(text) // chars_per_line) + 1
            cell_height = estimated_lines * 18
            needed_height = max(needed_height, cell_height)

        ws.row_dimensions[row_num].height = min(max(needed_height, 35), 409)

    ws.freeze_panes = "A2"

    if lock_sheet:
        ws.protection.sheet = True
        ws.protection.password = "J1825"

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


# =========================
# مطابقة الطالبات الذكية
# =========================

def normalize_arabic(text):
    """تنظيف النص العربي لتسهيل المطابقة"""
    if not text or pd.isna(text):
        return ""
    text = str(text).strip()
    # توحيد الأحرف العربية الشائعة
    text = re.sub(r'[أإآا]', 'ا', text)
    text = re.sub(r'[ةه]', 'ه', text)
    text = re.sub(r'[يىئ]', 'ي', text)
    text = re.sub(r'[ؤو]', 'و', text)
    # حذف التشكيل
    text = re.sub(r'[\u064B-\u065F]', '', text)
    # حذف المسافات الزائدة والمسافات والشرطات
    text = re.sub(r'[\s\-_]+', ' ', text).strip().lower()
    return text


def normalize_english(text):
    """تنظيف النص الإنجليزي لتسهيل المطابقة"""
    if not text or pd.isna(text):
        return ""
    text = str(text).strip().lower()
    text = re.sub(r'[\s\-_]+', ' ', text).strip()
    return text


def normalize_id(text):
    """تنظيف الرقم الأكاديمي"""
    if not text or pd.isna(text):
        return ""
    text = str(text).strip()
    # إزالة المسافات والشرطات
    text = re.sub(r'[\s\-_]+', '', text)
    return text.lower()


def normalize_email(text):
    """تنظيف الإيميل"""
    # إذا كان Series أو قيمة pandas، نحوله لـ scalar أولاً
    try:
        if hasattr(text, 'item'):
            text = text.item()
        elif hasattr(text, '__len__') and not isinstance(text, str):
            text = str(text)
    except Exception:
        pass
    if text is None:
        return ""
    try:
        if pd.isna(text):
            return ""
    except (TypeError, ValueError):
        pass
    return str(text).strip().lower()


def fuzzy_name_match(name1, name2, threshold=0.75):
    """مطابقة أسماء بنسبة تشابه — بدون مكتبات خارجية"""
    n1 = normalize_arabic(name1) if any('\u0600' <= c <= '\u06FF' for c in str(name1)) else normalize_english(name1)
    n2 = normalize_arabic(name2) if any('\u0600' <= c <= '\u06FF' for c in str(name2)) else normalize_english(name2)

    if not n1 or not n2:
        return False, 0.0

    if n1 == n2:
        return True, 1.0

    # مطابقة جزئية — هل كل كلمة في الاسم الأقصر موجودة في الأطول؟
    words1 = set(n1.split())
    words2 = set(n2.split())
    if words1 and words2:
        shorter = words1 if len(words1) <= len(words2) else words2
        longer = words2 if len(words1) <= len(words2) else words1
        match_count = sum(1 for w in shorter if any(w in lw or lw in w for lw in longer))
        ratio = match_count / max(len(shorter), 1)
        if ratio >= threshold:
            return True, ratio

    # Levenshtein بسيط
    len1, len2 = len(n1), len(n2)
    if max(len1, len2) == 0:
        return True, 1.0
    matrix = [[0] * (len2 + 1) for _ in range(len1 + 1)]
    for i in range(len1 + 1):
        matrix[i][0] = i
    for j in range(len2 + 1):
        matrix[0][j] = j
    for i in range(1, len1 + 1):
        for j in range(1, len2 + 1):
            cost = 0 if n1[i-1] == n2[j-1] else 1
            matrix[i][j] = min(matrix[i-1][j] + 1, matrix[i][j-1] + 1, matrix[i-1][j-1] + cost)
    distance = matrix[len1][len2]
    similarity = 1 - distance / max(len1, len2)
    return similarity >= threshold, similarity


def detect_roster_columns(df_roster):
    """اكتشاف أعمدة ملف المعنيات تلقائياً"""
    cols = {
        'email': None,
        'name_en': None,
        'name_ar': None,
        'student_id': None,
        'section': None,
    }
    for col in df_roster.columns:
        h = clean_header(col).lower()
        if any(x in h for x in ['email', 'إيميل', 'ايميل', 'بريد']):
            cols['email'] = col
        elif any(x in h for x in ['رقم أكاديمي', 'رقم الاكاديمي', 'الرقم الأكاديمي', 'academic', 'id', 'رقم']):
            cols['student_id'] = col
        elif any(x in h for x in ['شعبة', 'فصل', 'section', 'class']):
            cols['section'] = col
        elif any(x in h for x in ['اسم', 'name']) and any('\u0600' <= c <= '\u06FF' for c in col):
            cols['name_ar'] = col
        elif any(x in h for x in ['name', 'اسم']):
            if any('\u0600' <= c <= '\u06FF' for c in col):
                if not cols['name_ar']:
                    cols['name_ar'] = col
            else:
                if not cols['name_en']:
                    cols['name_en'] = col
    return cols


def detect_response_columns(df_responses):
    """اكتشاف أعمدة ملف الاستجابات"""
    cols = {
        'email': None,
        'name_en': None,
        'name_ar': None,
        'student_id': None,
    }
    for col in df_responses.columns:
        h = clean_header(col).lower()
        if h == 'email':
            cols['email'] = col
        elif h == 'name':
            cols['name_en'] = col
        elif any(x in h for x in ['الاسم الرباعي', 'اسم الطالبة', 'الاسم']):
            cols['name_ar'] = col
        elif any(x in h for x in ['الرقم الأكاديمي', 'رقم أكاديمي', 'رقم الاكاديمي']):
            cols['student_id'] = col
    return cols


def _row_val(row, col, default=''):
    """استخراج قيمة من صف DataFrame بأمان"""
    if not col:
        return default
    try:
        val = row[col] if col in row.index else default
        if pd.isna(val):
            return default
        return val
    except Exception:
        return default


def run_attendance_check(df_responses, df_roster, resp_cols, roster_cols):
    """
    فحص الحضور والغياب مع ثلاث حالات:
    1. إيميل صحيح + اسم صحيح  → مصرح ✅
    2. إيميل صحيح + اسم مختلف → مصرح لكن تحذير ⚠️ (name_mismatch)
    3. إيميل غير موجود في المعنيات → غير مصرح ❌
    """
    # بناء dict: إيميل → بيانات الطالبة من ملف المعنيات
    roster_by_email = {}
    if roster_cols['email']:
        for _, r in df_roster.iterrows():
            e = normalize_email(_row_val(r, roster_cols['email']))
            if e:
                roster_by_email[e] = r

    # فحص كل استجابة
    results = []
    for idx, row in df_responses.iterrows():
        email_val    = str(_row_val(row, resp_cols['email'])       if resp_cols['email']       else '').strip()
        name_en_val  = str(_row_val(row, resp_cols['name_en'])     if resp_cols['name_en']     else '').strip()
        name_ar_val  = str(_row_val(row, resp_cols['name_ar'])     if resp_cols['name_ar']     else '').strip()
        id_val       = str(_row_val(row, resp_cols['student_id'])  if resp_cols['student_id']  else '').strip()

        resp_email = normalize_email(email_val)
        name_mismatch = False
        expected_name = ''
        method = ''

        if roster_by_email:
            if resp_email in roster_by_email:
                matched = True
                roster_entry = roster_by_email[resp_email]

                # تحقق من الاسم الإنجليزي
                if resp_cols['name_en'] and roster_cols['name_en']:
                    roster_name_en = str(_row_val(roster_entry, roster_cols['name_en'])).strip()
                    ok, _ = fuzzy_name_match(name_en_val, roster_name_en)
                    if not ok and name_en_val and roster_name_en:
                        name_mismatch = True
                        expected_name = roster_name_en

                # تحقق من الاسم العربي (إضافي)
                if not name_mismatch and resp_cols['name_ar'] and roster_cols['name_ar']:
                    roster_name_ar = str(_row_val(roster_entry, roster_cols['name_ar'])).strip()
                    ok, _ = fuzzy_name_match(name_ar_val, roster_name_ar)
                    if not ok and name_ar_val and roster_name_ar:
                        name_mismatch = True
                        expected_name = roster_name_ar

                method = "إيميل ✅ — اسم مختلف ⚠️" if name_mismatch else "إيميل ✅"
            else:
                matched = False
                method = "إيميل غير مصرح ❌"
        else:
            matched = True
            method = "لا يوجد إيميلات للمقارنة"

        results.append({
            'row_index':     idx,
            'email':         email_val,
            'name_en':       name_en_val,
            'name_ar':       name_ar_val,
            'student_id':    id_val,
            'matched':       matched,
            'name_mismatch': name_mismatch,
            'expected_name': expected_name,
            'match_method':  method,
        })

    # الغائبات — في المعنيات لكن إيميلهم ما ظهر في الاستجابات
    submitted_email_set = set()
    if resp_cols['email']:
        for _, row in df_responses.iterrows():
            e = normalize_email(_row_val(row, resp_cols['email']))
            if e:
                submitted_email_set.add(e)

    absent = []
    for _, roster_row in df_roster.iterrows():
        r_email = normalize_email(_row_val(roster_row, roster_cols['email'])) if roster_cols['email'] else ''
        if r_email and r_email not in submitted_email_set:
            absent.append({
                'الإيميل':        _row_val(roster_row, roster_cols['email'])       if roster_cols['email']       else '',
                'الاسم العربي':   _row_val(roster_row, roster_cols['name_ar'])     if roster_cols['name_ar']     else '',
                'الاسم الإنجليزي': _row_val(roster_row, roster_cols['name_en'])   if roster_cols['name_en']     else '',
                'الرقم الأكاديمي': _row_val(roster_row, roster_cols['student_id']) if roster_cols['student_id'] else '',
                'الشعبة':         _row_val(roster_row, roster_cols['section'])     if roster_cols['section']     else '',
            })

    unauthorized   = [r for r in results if not r['matched']]
    name_mismatches = [r for r in results if r['matched'] and r['name_mismatch']]
    return results, absent, unauthorized, name_mismatches


def highlight_unauthorized_excel(df_responses, unauthorized_indices, color="FF9999"):
    """إنشاء ملف Excel مع تلوين صفوف محددة بلون قابل للتخصيص"""
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df_responses.to_excel(writer, index=False, sheet_name="الاستجابات")

    output.seek(0)
    wb = load_workbook(output)
    ws = wb.active

    red_fill = PatternFill("solid", fgColor=color)

    # رأس الجدول
    for cell in ws[1]:
        cell.fill = PatternFill("solid", fgColor="BFEFFF")
        cell.font = Font(bold=True, size=12)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    highlighted_set = set(unauthorized_indices)

    for row_num, row in enumerate(ws.iter_rows(min_row=2), start=2):
        df_row_idx = row_num - 2
        is_highlighted = df_row_idx in highlighted_set
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.font = Font(size=16)
            if is_highlighted:
                cell.fill = red_fill

    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 45

    final_output = BytesIO()
    wb.save(final_output)
    final_output.seek(0)
    return final_output




def get_single_school_name_from_df(df, school_col=None):
    """يرجع اسم المدرسة إذا كان الملف يخص مدرسة واحدة، وإلا يرجع متعدد المدارس."""
    if school_col and school_col in df.columns:
        values = [str(v).strip() for v in df[school_col].dropna().unique() if str(v).strip()]
        if len(values) == 1:
            return values[0]
        if len(values) > 1:
            return "متعدد المدارس"
    return "بدون مدرسة"


def format_simple_table_excel(df_table, sheet_name="البيانات", highlight_total=False):
    """تنسيق عام لملفات Excel الصغيرة مثل الغياب والرصد."""
    buf = BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as wr:
        df_table.to_excel(wr, index=False, sheet_name=sheet_name)
    buf.seek(0)

    wb = load_workbook(buf)
    ws = wb.active
    ws.sheet_view.rightToLeft = True

    hdr_fill  = PatternFill("solid", fgColor="15396B")
    alt_fill  = PatternFill("solid", fgColor="EAF6FF")
    white_fill = PatternFill("solid", fgColor="FFFFFF")
    tot_fill  = PatternFill("solid", fgColor="FFF2CC")
    thin_b    = Border(
        left=Side(style="thin", color="AAAAAA"),
        right=Side(style="thin", color="AAAAAA"),
        top=Side(style="thin", color="AAAAAA"),
        bottom=Side(style="thin", color="AAAAAA"),
    )
    center_al = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for cell in ws[1]:
        cell.fill      = hdr_fill
        cell.font      = Font(bold=True, size=13, color="FFFFFF")
        cell.alignment = center_al
        cell.border    = thin_b
    ws.row_dimensions[1].height = 40

    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        bg = alt_fill if row_idx % 2 == 0 else white_fill
        for cell in row:
            header_val = clean_header(ws.cell(1, cell.column).value).lower()
            is_total = highlight_total and ("total" in header_val or "points" in header_val or "الدرجة" in header_val)
            cell.fill      = tot_fill if is_total else bg
            cell.font      = Font(size=12, bold=is_total)
            cell.alignment = center_al
            cell.border    = thin_b
        ws.row_dimensions[row_idx].height = 28

    for col in ws.columns:
        col_letter = col[0].column_letter
        max_len = 0
        for cell in col:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        if max_len <= 15:
            width = 18
        elif max_len <= 35:
            width = 28
        elif max_len <= 70:
            width = 38
        else:
            width = 50
        ws.column_dimensions[col_letter].width = width

    ws.freeze_panes = "A2"

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return out


# =========================
# ترتيب ملفات التجميع
# =========================
def get_part_number(filename):
    match = re.search(r"(?:مظروف|جزء|[-_\s])(\d+)(?:\.xlsx)?$", filename)
    if match:
        return int(match.group(1))

    nums = re.findall(r"\d+", filename)
    if nums:
        return int(nums[-1])

    return 999999


def detect_base_name_from_parts(files):
    if not files:
        return "الاستجابات"

    first_name = Path(files[0].name).stem
    base = re.sub(r"[-_\s]*(\d+)$", "", first_name).strip()
    base = re.sub(r"^(مظروف|جزء)[-_\s]*", "", base).strip()
    return safe_filename(base or "الاستجابات")


# =========================
# حساب الدرجات النهائية للرصد والتجميع
# =========================
def get_score_columns_for_total(df):
    """يرجع أعمدة الدرجات الحقيقية فقط، ويستثني Total Points ودرجات بيانات الطالبة."""
    score_cols = []
    for col in df.columns:
        header = clean_header(col)
        h = header.lower()
        original_header = header.split(" / الدرجة من ")[0].strip()
        if (
            "points" in h
            and h != "total points"
            and not _is_student_info_field(original_header)
        ):
            score_cols.append(col)
    return score_cols


def recompute_total_points_dataframe(df):
    """يعيد حساب Total Points من أعمدة Points داخل DataFrame قبل الرصد أو التجميع."""
    df = df.copy()
    score_cols = get_score_columns_for_total(df)

    if not score_cols:
        return df, None, []

    total_values = pd.Series(0.0, index=df.index)
    for col in score_cols:
        total_values = total_values + pd.to_numeric(df[col], errors="coerce").fillna(0)

    total_values = total_values.round(2)
    total_values = total_values.apply(lambda x: int(x) if float(x).is_integer() else x)

    total_col = None
    for col in df.columns:
        if clean_header(col).lower() == "total points":
            total_col = col
            break

    if total_col is None:
        total_col = "Total Points"
        df[total_col] = total_values
    else:
        df[total_col] = total_values

    return df, total_col, score_cols


# =========================
# تصميم الواجهة الاحترافي
# =========================
def apply_ui_style():
    logo_base64 = get_logo_base64()
    logo_html = ""
    if logo_base64:
        logo_html = f'<img src="data:image/png;base64,{logo_base64}" class="logo-img">'

    st.markdown(f"""
    <link href="https://fonts.googleapis.com/css2?family=Cairo:wght@300;400;600;700;900&family=Tajawal:wght@300;400;500;700;800&display=swap" rel="stylesheet">

    <style>
    /* ===== المتغيرات الأساسية ===== */
    :root {{
        --primary:       #0d3060;
        --primary-light: #1a4f8a;
        --accent:        #c8a035;
        --accent-light:  #e8c060;
        --surface:       #f4f7fc;
        --surface-card:  #ffffff;
        --border:        #d6e4f7;
        --text-main:     #1a2a45;
        --text-muted:    #5a7092;
        --success:       #1e8c5a;
        --warning:       #c47a1a;
        --danger:        #c0392b;
        --info:          #1565a0;
        --shadow-sm:     0 2px 8px rgba(13,48,96,0.08);
        --shadow-md:     0 6px 24px rgba(13,48,96,0.12);
        --shadow-lg:     0 12px 40px rgba(13,48,96,0.18);
        --radius:        14px;
        --radius-sm:     8px;
    }}

    /* ===== القاعدة ===== */
    html, body, [class*="css"] {{
        direction: rtl;
        text-align: right;
        font-family: 'Cairo', 'Tajawal', sans-serif !important;
        font-size: 16px !important;
        color: var(--text-main);
        background-color: var(--surface) !important;
    }}

    /* ===== الحاوية الرئيسية ===== */
    .block-container {{
        max-width: 1300px !important;
        margin: auto;
        padding: 0 2rem 4rem 2rem !important;
        background: var(--surface);
    }}

    /* ===== شريط العنوان الرئيسي ===== */
    .hero-banner {{
        background: linear-gradient(135deg, var(--primary) 0%, var(--primary-light) 55%, #1e6abf 100%);
        border-radius: 0 0 var(--radius) var(--radius);
        padding: 36px 40px 32px;
        margin: -1rem -2rem 2rem -2rem;
        display: flex;
        align-items: center;
        gap: 24px;
        position: relative;
        overflow: hidden;
        box-shadow: var(--shadow-lg);
    }}

    .hero-banner::before {{
        content: '';
        position: absolute;
        top: -60px; left: -60px;
        width: 260px; height: 260px;
        background: rgba(200,160,53,0.12);
        border-radius: 50%;
        pointer-events: none;
    }}

    .hero-banner::after {{
        content: '';
        position: absolute;
        bottom: -80px; right: -40px;
        width: 320px; height: 320px;
        background: rgba(255,255,255,0.04);
        border-radius: 50%;
        pointer-events: none;
    }}

    .logo-img {{
        height: 80px;
        width: auto;
        filter: drop-shadow(0 2px 8px rgba(0,0,0,0.25));
        flex-shrink: 0;
    }}

    .hero-text {{
        flex: 1;
        position: relative;
        z-index: 2;
    }}

    .hero-text h1 {{
        color: #ffffff;
        font-family: 'Cairo', sans-serif;
        font-size: 28px;
        font-weight: 900;
        margin: 0 0 6px 0;
        line-height: 1.4;
        letter-spacing: -0.3px;
        text-shadow: 0 2px 8px rgba(0,0,0,0.2);
    }}

    .hero-subtitle {{
        color: var(--accent-light);
        font-size: 14px;
        font-weight: 500;
        letter-spacing: 0.5px;
        opacity: 0.9;
    }}

    .hero-badge {{
        background: rgba(200,160,53,0.2);
        border: 1px solid rgba(200,160,53,0.4);
        color: var(--accent-light);
        font-size: 12px;
        font-weight: 600;
        padding: 4px 12px;
        border-radius: 20px;
        display: inline-block;
        margin-top: 8px;
    }}

    /* ===== شريط معلومات المستخدم ===== */
    .user-info-bar {{
        background: var(--surface-card);
        border: 1px solid var(--border);
        border-radius: var(--radius-sm);
        padding: 10px 18px;
        display: flex;
        align-items: center;
        gap: 10px;
        font-size: 14px;
        font-weight: 600;
        color: var(--primary);
        box-shadow: var(--shadow-sm);
        margin-bottom: 1rem;
    }}

    /* ===== الأزرار ===== */
    .stButton > button {{
        font-family: 'Cairo', sans-serif !important;
        font-size: 15px !important;
        font-weight: 700 !important;
        padding: 10px 24px !important;
        border-radius: var(--radius-sm) !important;
        transition: all 0.22s ease !important;
        border: none !important;
        letter-spacing: 0.2px;
    }}

    .stButton > button[kind="primary"] {{
        background: linear-gradient(135deg, var(--primary) 0%, var(--primary-light) 100%) !important;
        color: #fff !important;
        box-shadow: 0 4px 14px rgba(13,48,96,0.28) !important;
    }}

    .stButton > button[kind="primary"]:hover {{
        transform: translateY(-2px) !important;
        box-shadow: 0 8px 22px rgba(13,48,96,0.36) !important;
    }}

    .stButton > button:not([kind="primary"]) {{
        background: var(--surface-card) !important;
        color: var(--primary) !important;
        border: 1.5px solid var(--border) !important;
        box-shadow: var(--shadow-sm) !important;
    }}

    .stButton > button:not([kind="primary"]):hover {{
        background: var(--surface) !important;
        border-color: var(--primary) !important;
        transform: translateY(-1px) !important;
    }}

    /* ===== التبويبات ===== */
    .stTabs [data-baseweb="tab-list"] {{
        background: var(--surface-card);
        border-radius: var(--radius-sm);
        padding: 6px;
        gap: 6px;
        border: 1px solid var(--border);
        box-shadow: var(--shadow-sm);
        margin-bottom: 1.5rem;
    }}

    .stTabs [data-baseweb="tab"] {{
        font-family: 'Cairo', sans-serif !important;
        font-size: 15px !important;
        font-weight: 700 !important;
        color: var(--text-muted) !important;
        border-radius: var(--radius-sm) !important;
        padding: 10px 22px !important;
        transition: all 0.2s ease !important;
        background: transparent !important;
    }}

    .stTabs [aria-selected="true"] {{
        background: linear-gradient(135deg, var(--primary) 0%, var(--primary-light) 100%) !important;
        color: #ffffff !important;
        box-shadow: 0 4px 12px rgba(13,48,96,0.25) !important;
    }}

    /* ===== حقول الإدخال ===== */
    .stTextInput > div > div > input,
    .stNumberInput > div > div > input {{
        font-family: 'Cairo', sans-serif !important;
        font-size: 15px !important;
        border: 1.5px solid var(--border) !important;
        border-radius: var(--radius-sm) !important;
        padding: 10px 14px !important;
        background: var(--surface-card) !important;
        color: var(--text-main) !important;
        transition: border-color 0.2s !important;
        direction: rtl !important;
    }}

    .stTextInput > div > div > input:focus,
    .stNumberInput > div > div > input:focus {{
        border-color: var(--primary) !important;
        box-shadow: 0 0 0 3px rgba(13,48,96,0.1) !important;
    }}

    /* ===== القوائم المنسدلة ===== */
    .stSelectbox > div > div,
    .stMultiSelect > div > div {{
        font-family: 'Cairo', sans-serif !important;
        font-size: 15px !important;
        border: 1.5px solid var(--border) !important;
        border-radius: var(--radius-sm) !important;
        background: var(--surface-card) !important;
    }}

    /* ===== رفع الملفات ===== */
    .stFileUploader > div {{
        border: 2px dashed var(--border) !important;
        border-radius: var(--radius) !important;
        background: linear-gradient(135deg, #f0f6ff 0%, #f8fbff 100%) !important;
        padding: 24px !important;
        text-align: center !important;
        transition: all 0.25s ease !important;
    }}

    .stFileUploader > div:hover {{
        border-color: var(--primary) !important;
        background: linear-gradient(135deg, #e8f2ff 0%, #f4f9ff 100%) !important;
        box-shadow: var(--shadow-sm) !important;
    }}

    .stFileUploader label {{
        font-family: 'Cairo', sans-serif !important;
        font-size: 15px !important;
        font-weight: 600 !important;
        color: var(--primary) !important;
    }}

    /* ===== رسائل التنبيه ===== */
    .stSuccess, .stInfo, .stWarning, .stError {{
        border-radius: var(--radius-sm) !important;
        font-family: 'Cairo', sans-serif !important;
        font-size: 14px !important;
        font-weight: 600 !important;
        border: none !important;
        padding: 12px 18px !important;
        box-shadow: var(--shadow-sm) !important;
    }}

    [data-testid="stNotification"] {{
        border-radius: var(--radius-sm) !important;
        font-family: 'Cairo', sans-serif !important;
        font-weight: 600 !important;
    }}

    div[data-testid="stSuccess"] {{
        background: linear-gradient(135deg, #e8f8f0 0%, #f0faf5 100%) !important;
        border-right: 4px solid var(--success) !important;
        color: #0f5235 !important;
    }}

    div[data-testid="stInfo"] {{
        background: linear-gradient(135deg, #e8f2ff 0%, #f0f7ff 100%) !important;
        border-right: 4px solid var(--info) !important;
        color: #0d3060 !important;
    }}

    div[data-testid="stWarning"] {{
        background: linear-gradient(135deg, #fff8e8 0%, #fffcf0 100%) !important;
        border-right: 4px solid var(--warning) !important;
        color: #7a4a0a !important;
    }}

    div[data-testid="stError"] {{
        background: linear-gradient(135deg, #ffeaea 0%, #fff5f5 100%) !important;
        border-right: 4px solid var(--danger) !important;
        color: #7a1a1a !important;
    }}

    /* ===== الـ Dataframe ===== */
    .stDataFrame {{
        border-radius: var(--radius-sm) !important;
        overflow: hidden !important;
        box-shadow: var(--shadow-sm) !important;
        border: 1px solid var(--border) !important;
    }}

    /* ===== الـ Expander ===== */
    .streamlit-expanderHeader {{
        font-family: 'Cairo', sans-serif !important;
        font-size: 15px !important;
        font-weight: 700 !important;
        color: var(--primary) !important;
        background: var(--surface-card) !important;
        border-radius: var(--radius-sm) !important;
        padding: 12px 18px !important;
        border: 1px solid var(--border) !important;
        transition: all 0.2s !important;
    }}

    .streamlit-expanderHeader:hover {{
        background: var(--surface) !important;
        border-color: var(--primary) !important;
    }}

    .streamlit-expanderContent {{
        border: 1px solid var(--border) !important;
        border-top: none !important;
        border-radius: 0 0 var(--radius-sm) var(--radius-sm) !important;
        padding: 16px !important;
        background: var(--surface-card) !important;
    }}

    /* ===== Checkbox & Radio ===== */
    .stCheckbox label,
    .stRadio label {{
        font-family: 'Cairo', sans-serif !important;
        font-size: 15px !important;
        font-weight: 600 !important;
        color: var(--text-main) !important;
    }}

    /* ===== العناوين الفرعية ===== */
    [data-testid="stMarkdownContainer"] h3 {{
        font-family: 'Cairo', sans-serif !important;
        font-size: 20px !important;
        font-weight: 800 !important;
        color: var(--primary) !important;
        text-align: center !important;
        padding: 14px 20px !important;
        background: linear-gradient(135deg, #eaf3ff 0%, #f4f8ff 100%);
        border-radius: var(--radius-sm);
        border-right: 5px solid var(--accent);
        margin: 20px 0 16px 0 !important;
        box-shadow: var(--shadow-sm);
    }}

    [data-testid="stMarkdownContainer"] h4 {{
        font-family: 'Cairo', sans-serif !important;
        font-size: 17px !important;
        font-weight: 700 !important;
        color: var(--primary) !important;
    }}

    /* ===== الـ Container / Border ===== */
    [data-testid="stVerticalBlockBorderWrapper"] > div {{
        border-radius: var(--radius) !important;
        border: 1px solid var(--border) !important;
        padding: 18px 20px !important;
        background: var(--surface-card) !important;
        box-shadow: var(--shadow-sm) !important;
        margin-bottom: 12px !important;
        transition: box-shadow 0.2s !important;
    }}

    [data-testid="stVerticalBlockBorderWrapper"] > div:hover {{
        box-shadow: var(--shadow-md) !important;
    }}

    /* ===== تقسيم ===== */
    hr {{
        border: none !important;
        border-top: 2px solid var(--border) !important;
        margin: 28px 0 !important;
        opacity: 0.6 !important;
    }}

    /* ===== Metrics ===== */
    [data-testid="metric-container"] {{
        background: var(--surface-card) !important;
        border-radius: var(--radius-sm) !important;
        border: 1px solid var(--border) !important;
        padding: 16px !important;
        box-shadow: var(--shadow-sm) !important;
        text-align: center !important;
        transition: transform 0.2s, box-shadow 0.2s !important;
    }}

    [data-testid="metric-container"]:hover {{
        transform: translateY(-2px) !important;
        box-shadow: var(--shadow-md) !important;
    }}

    [data-testid="metric-container"] [data-testid="stMetricLabel"] {{
        font-family: 'Cairo', sans-serif !important;
        font-size: 13px !important;
        font-weight: 700 !important;
        color: var(--text-muted) !important;
    }}

    [data-testid="metric-container"] [data-testid="stMetricValue"] {{
        font-family: 'Tajawal', sans-serif !important;
        font-size: 28px !important;
        font-weight: 800 !important;
        color: var(--primary) !important;
    }}

    /* ===== Labels عامة ===== */
    label {{
        font-family: 'Cairo', sans-serif !important;
        font-size: 15px !important;
        font-weight: 700 !important;
        color: var(--text-main) !important;
    }}

    /* ===== الـ Spinner ===== */
    .stSpinner {{
        text-align: center;
        color: var(--primary) !important;
    }}

    /* ===== الـ Download Button ===== */
    .stDownloadButton > button {{
        font-family: 'Cairo', sans-serif !important;
        font-size: 14px !important;
        font-weight: 700 !important;
        background: linear-gradient(135deg, var(--accent) 0%, #d4a830 100%) !important;
        color: var(--primary) !important;
        border: none !important;
        border-radius: var(--radius-sm) !important;
        padding: 10px 22px !important;
        box-shadow: 0 4px 14px rgba(200,160,53,0.3) !important;
        transition: all 0.22s ease !important;
    }}

    .stDownloadButton > button:hover {{
        transform: translateY(-2px) !important;
        box-shadow: 0 8px 22px rgba(200,160,53,0.4) !important;
    }}

    /* ===== الـ Info Badge (اسم المستخدم) ===== */
    [data-testid="stAlert"] {{
        border-radius: var(--radius-sm) !important;
        font-family: 'Cairo', sans-serif !important;
        font-size: 14px !important;
        font-weight: 600 !important;
    }}

    /* ===== الـ Caption ===== */
    .stCaption, small {{
        font-family: 'Cairo', sans-serif !important;
        font-size: 12px !important;
        color: var(--text-muted) !important;
    }}

    /* ===== Sidebar (if any) ===== */
    .css-1d391kg {{
        background: var(--primary) !important;
    }}

    /* ===== تمييز الصفحة للـ RTL ===== */
    .stApp {{
        direction: rtl;
    }}

    </style>

    <!-- ===== Hero Banner ===== -->
    <div class="hero-banner">
        {logo_html}
        <div class="hero-text">
            <h1>مركز التصحيح المركزي</h1>
            <div class="hero-subtitle">مدرسة جدحفص الثانوية للبنات</div>
            <div class="hero-badge">📅 {today}</div>
        </div>
    </div>
    """, unsafe_allow_html=True)


# =========================
# التطبيق
# =========================
if not login_screen():
    st.stop()

# تهيئة session_state للمطابقة
if "attendance_done" not in st.session_state:
    st.session_state["attendance_done"] = False
if "attendance_results" not in st.session_state:
    st.session_state["attendance_results"] = []
if "absent_list" not in st.session_state:
    st.session_state["absent_list"] = []
if "unauthorized_list" not in st.session_state:
    st.session_state["unauthorized_list"] = []

apply_ui_style()

col_user, col_logout = st.columns([4, 1])
with col_user:
    st.info(f"مرحبًا، {st.session_state.username} — نوع الدخول: {st.session_state.role}")
with col_logout:
    if st.button("تسجيل خروج"):
        st.session_state.logged_in = False
        st.session_state.role = None
        st.session_state.username = None
        st.rerun()

if st.session_state.role == "admin":
    tab1, tab2, tab3 = st.tabs(["✂️ تقسيم الاستجابات", "📥 تجميع الاستجابات", "📊 تقرير الأدمن"])
else:
    tab1, tab2 = st.tabs(["✂️ تقسيم الاستجابات", "📥 تجميع الاستجابات"])
    tab3 = None


# =========================
# تبويب التقسيم
# =========================
with tab1:
    st.markdown("<h3>✂️ تقسيم ملف الاستجابات</h3>", unsafe_allow_html=True)

    uploaded_file = st.file_uploader(
        "ارفع ملف Excel الأصلي للاستجابات",
        type=["xlsx"],
        key="split_file",
    )

    if uploaded_file:
        df = pd.read_excel(uploaded_file, engine="openpyxl")
        original_file_name = uploaded_file.name

        # =========================
        # تحديد مدرسة الاستجابات
        # =========================
        school_col = detect_school_column(df)
        if school_col:
            st.success(f"✅ تم اكتشاف عمود المدرسة في الملف: {school_col}")
        else:
            st.warning("⚠️ لا يوجد عمود مدرسة في ملف الاستجابات.")

            skip_school = st.checkbox(
                "⏭️ تخطي تحديد اسم المدرسة والانتقال للخطوة التالية",
                key=f"skip_school_{original_file_name}",
                help="استخدمي هذا الخيار إذا كان الامتحان لا يحتاج رصد المدرسة أو سيتم التعامل معها لاحقًا.",
            )

            if skip_school:
                school_col = None
                st.info("تم تخطي خطوة تحديد المدرسة بإقرار منك. لن يضيف البرنامج عمود المدرسة لهذا الملف.")
            else:
                selected_school = st.selectbox(
                    "اختاري المدرسة صاحبة ملف الاستجابات",
                    SCHOOL_OPTIONS,
                    key=f"school_select_{original_file_name}",
                )
                df["المدرسة"] = selected_school
                school_col = "المدرسة"
                st.info(f"تمت إضافة عمود المدرسة لكل الاستجابات: {selected_school} — وسيكون مخفيًا عن المصححات.")

        st.success(f"✅ تم رفع ملف الاستجابات بنجاح — عدد الاستجابات: {len(df)}")

        # =========================
        # خطوة 1: مطابقة قائمة المعنيات
        # =========================
        st.markdown("---")
        st.markdown("### 📋 خطوة 1: مطابقة قائمة الطالبات المعنيات بالامتحان")

        roster_file = st.file_uploader(
            "ارفعي ملف Excel لقائمة الطالبات المعنيات بهذا الامتحان",
            type=["xlsx"],
            key="roster_file",
            help="يجب أن يحتوي على أعمدة: الإيميل، الاسم، الرقم الأكاديمي (أو أي منها). سيكتشف البرنامج الأعمدة تلقائياً.",
        )

        attendance_passed = False  # نعرف إذا اجتازت خطوة المطابقة

        if roster_file:
            df_roster = pd.read_excel(roster_file, engine="openpyxl")
            st.success(f"✅ تم رفع ملف المعنيات — عدد الطالبات: {len(df_roster)}")

            # اكتشاف الأعمدة تلقائياً — نحفظها في session_state حتى لا تُعاد عند كل rerun
            sig_key = f"cols_detected_{original_file_name}_{roster_file.name}"
            if sig_key not in st.session_state:
                st.session_state[sig_key] = {
                    'resp': detect_response_columns(df),
                    'roster': detect_roster_columns(df_roster),
                }

            detected_resp   = st.session_state[sig_key]['resp']
            detected_roster = st.session_state[sig_key]['roster']

            all_resp_cols   = ["— لا يوجد —"] + list(df.columns)
            all_roster_cols = ["— لا يوجد —"] + list(df_roster.columns)

            def _idx(lst, val):
                return lst.index(val) if val in lst else 0

            with st.expander("🔍 مراجعة الأعمدة المكتشفة (اضغطي للتعديل إذا لزم)", expanded=False):
                col_r1, col_r2 = st.columns(2)
                with col_r1:
                    st.markdown("**ملف الاستجابات:**")
                    st.selectbox("عمود الإيميل",         all_resp_cols, index=_idx(all_resp_cols, detected_resp.get('email')),      key="resp_email_col")
                    st.selectbox("عمود الاسم الإنجليزي", all_resp_cols, index=_idx(all_resp_cols, detected_resp.get('name_en')),    key="resp_nameen_col")
                    st.selectbox("عمود الاسم العربي",     all_resp_cols, index=_idx(all_resp_cols, detected_resp.get('name_ar')),    key="resp_namear_col")
                    st.selectbox("عمود الرقم الأكاديمي", all_resp_cols, index=_idx(all_resp_cols, detected_resp.get('student_id')), key="resp_id_col")
                with col_r2:
                    st.markdown("**ملف المعنيات:**")
                    st.selectbox("عمود الإيميل",         all_roster_cols, index=_idx(all_roster_cols, detected_roster.get('email')),      key="roster_email_col")
                    st.selectbox("عمود الاسم الإنجليزي", all_roster_cols, index=_idx(all_roster_cols, detected_roster.get('name_en')),    key="roster_nameen_col")
                    st.selectbox("عمود الاسم العربي",     all_roster_cols, index=_idx(all_roster_cols, detected_roster.get('name_ar')),    key="roster_namear_col")
                    st.selectbox("عمود الرقم الأكاديمي", all_roster_cols, index=_idx(all_roster_cols, detected_roster.get('student_id')), key="roster_id_col")
                    st.selectbox("عمود الشعبة (اختياري)", all_roster_cols, index=_idx(all_roster_cols, detected_roster.get('section')),   key="roster_section_col")

            def _none(k, state_key):
                v = st.session_state.get(state_key, "— لا يوجد —")
                return None if (not v or v == "— لا يوجد —") else v

            resp_cols = {
                'email':      _none('email',      'resp_email_col'),
                'name_en':    _none('name_en',    'resp_nameen_col'),
                'name_ar':    _none('name_ar',    'resp_namear_col'),
                'student_id': _none('student_id', 'resp_id_col'),
            }
            roster_cols = {
                'email':      _none('email',      'roster_email_col'),
                'name_en':    _none('name_en',    'roster_nameen_col'),
                'name_ar':    _none('name_ar',    'roster_namear_col'),
                'student_id': _none('student_id', 'roster_id_col'),
                'section':    _none('section',    'roster_section_col'),
            }

            # إحصائية ملف المعنيات
            roster_total = len(df_roster)
            st.info(f"📋 عدد الطالبات في قائمة المعنيات: **{roster_total}** طالبة")

            if st.button("🔍 تشغيل فحص الحضور والمطابقة", type="primary"):
                with st.spinner("جاري المطابقة..."):
                    results, absent_list, unauthorized_list, name_mismatch_list = run_attendance_check(
                        df, df_roster, resp_cols, roster_cols
                    )
                    st.session_state["attendance_results"] = results
                    st.session_state["absent_list"] = absent_list
                    st.session_state["unauthorized_list"] = unauthorized_list
                    st.session_state["name_mismatch_list"] = name_mismatch_list
                    st.session_state["attendance_done"] = True

            if st.session_state.get("attendance_done"):
                results = st.session_state["attendance_results"]
                absent_list = st.session_state["absent_list"]
                unauthorized_list = st.session_state["unauthorized_list"]
                name_mismatch_list = st.session_state.get("name_mismatch_list", [])

                total = len(results)
                matched_count = sum(1 for r in results if r['matched'])
                unauth_count = len(unauthorized_list)
                absent_count = len(absent_list)
                mismatch_count = len(name_mismatch_list)

                # ملخص
                c1, c2, c3, c4, c5 = st.columns(5)
                c1.metric("📩 إجمالي الاستجابات", total)
                c2.metric("✅ مصرح واسم صحيح", matched_count - mismatch_count)
                c3.metric("⚠️ إيميل صح / اسم مختلف", mismatch_count)
                c4.metric("🚫 غير مصرح لهم", unauth_count)
                c5.metric("❌ غائبات", absent_count)

                # --- إيميل صحيح لكن اسم مختلف ---
                if name_mismatch_list:
                    st.warning(f"⚠️ يوجد {mismatch_count} طالبة إيميلها صحيح لكن الاسم المكتوب لا يطابق قائمة المعنيات — راجعيها:")
                    mismatch_df = pd.DataFrame([{
                        "الإيميل": r['email'],
                        "الاسم المكتوب في الاستجابة": r['name_en'] or r['name_ar'],
                        "الاسم المتوقع من القائمة": r['expected_name'],
                        "طريقة التحقق": r['match_method'],
                    } for r in name_mismatch_list])
                    st.dataframe(mismatch_df, use_container_width=True)

                    mismatch_indices = [r['row_index'] for r in name_mismatch_list]
                    mismatch_excel = highlight_unauthorized_excel(df, mismatch_indices, color="FFEB99")
                    st.download_button(
                        label="⬇️ تنزيل ملف الاستجابات مع تحديد الأسماء المختلفة (باللون الأصفر)",
                        data=mismatch_excel,
                        file_name=f"{safe_filename(Path(original_file_name).stem)}-أسماء-مختلفة.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key="dl_mismatch",
                    )
                else:
                    st.success("✅ جميع الأسماء متطابقة مع قائمة المعنيات.")

                # --- غير مصرح لهم ---
                if unauthorized_list:
                    st.error(f"🚫 تنبيه: يوجد {unauth_count} استجابة من طالبات غير مدرجات في قائمة المعنيات!")
                    unauth_df = pd.DataFrame([{
                        "الإيميل": r['email'],
                        "الاسم الإنجليزي": r['name_en'],
                        "الاسم العربي": r['name_ar'],
                        "الرقم الأكاديمي": r['student_id'],
                    } for r in unauthorized_list])
                    st.dataframe(unauth_df, use_container_width=True)

                    unauthorized_indices = [r['row_index'] for r in unauthorized_list]
                    colored_excel = highlight_unauthorized_excel(df, unauthorized_indices, color="FF9999")
                    st.download_button(
                        label="⬇️ تنزيل ملف الاستجابات مع تحديد الغير مصرح لهم (باللون الأحمر)",
                        data=colored_excel,
                        file_name=f"{safe_filename(Path(original_file_name).stem)}-مراجعة-المصرح-لهم.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key="dl_unauth",
                    )
                else:
                    st.success("✅ جميع من قدموا الامتحان مصرح لهم — لا يوجد إيميل غير مصرح له.")

                # --- الغياب ---
                if absent_list:
                    st.warning(f"❌ يوجد {absent_count} طالبة غائبة عن الامتحان:")
                    absent_df = pd.DataFrame(absent_list)
                    st.dataframe(absent_df, use_container_width=True)

                    # إضافة المدرسة لقائمة الغياب إذا لم تكن موجودة في القائمة الرسمية
                    absent_school_col = detect_school_column(absent_df)
                    school_name_for_file = get_single_school_name_from_df(df, school_col)
                    if not absent_school_col:
                        absent_df.insert(0, "المدرسة", school_name_for_file)

                    absent_buffer = format_simple_table_excel(absent_df, sheet_name="الغياب")
                    st.download_button(
                        label="⬇️ تنزيل قائمة الغياب Excel",
                        data=absent_buffer,
                        file_name=f"{safe_filename(school_name_for_file)}-قائمة-الغياب.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    )
                else:
                    st.success("✅ لا يوجد غياب — جميع الطالبات المعنيات قدّمن الامتحان.")

                # هل نكمل للتقسيم؟
                if unauth_count == 0 and absent_count == 0 and mismatch_count == 0:
                    st.success("🎉 المطابقة مكتملة بدون أي مشاكل. يمكنك الانتقال لخطوة التقسيم أدناه.")
                    attendance_passed = True
                else:
                    st.info("📌 راجعي النتائج أعلاه. يمكنك المتابعة للتقسيم عبر الزر أدناه.")
                    if st.checkbox("✅ راجعت النتائج وأريد المتابعة لخطوة التقسيم", key="confirm_proceed"):
                        attendance_passed = True
        else:
            # لو ما رفعت ملف المعنيات، تخطي الخطوة
            st.info("💡 يمكنك تخطي هذه الخطوة والانتقال للتقسيم مباشرة.")
            if st.checkbox("⏩ تخطي فحص المعنيات والانتقال للتقسيم مباشرة", key="skip_roster"):
                attendance_passed = True

        if attendance_passed:
          st.markdown("---")
          st.markdown("### ✂️ خطوة 2: إعدادات التقسيم")

          default_new_name = safe_filename(Path(original_file_name).stem)
          new_base_name = st.text_input(
              "اكتب اسم الملف الجديد قبل التقسيم",
              value=default_new_name,
              help="مثال: تقن106 — ستكون الملفات: تقن106-1، تقن106-2 ...",
          )
          new_base_name = safe_filename(new_base_name)

          chunk_size = st.number_input(
              "عدد الاستجابات في كل ملف",
              min_value=1,
              value=10,
              step=1,
          )

          all_columns = list(df.columns)
          auto_hidden_columns = [col for col in all_columns if should_auto_hide(col)]

          extra_hidden_columns = st.multiselect(
              "اختاري أي أعمدة إضافية تريدين إخفاءها قبل تنزيل الملفات",
              options=all_columns,
              default=[],
          )

          hidden_preview_columns = set(auto_hidden_columns + extra_hidden_columns)

          preview_chunk = df.iloc[0:chunk_size].copy()
          preview_chunk.insert(0, "رقم", range(1, len(preview_chunk) + 1))

          visible_preview_columns = [
              col for col in preview_chunk.columns
              if col not in hidden_preview_columns
          ]

          st.markdown("<h4 style='text-align:center;'>👁️ معاينة أول ملف بعد الإخفاء</h4>", unsafe_allow_html=True)
          st.dataframe(preview_chunk[visible_preview_columns], use_container_width=True)

          if auto_hidden_columns:
              st.info("الأعمدة التي سيخفيها التطبيق تلقائيًا: " + "، ".join([str(c) for c in auto_hidden_columns]))

          st.markdown("---")
          st.markdown("### 🟩 تحديد درجات الأسئلة")

          # لا نعتمد على أعلى درجة في الاستجابات كدرجة كبرى؛ لأنها قد تكون درجة طالبة فقط.
          # نحاول قراءة الدرجة الظاهرة في نص السؤال فقط، والناقص يترك 0.00 ليدخله المستخدم من الإجابة النموذجية.
          detected_scores, unknown_points = detect_max_scores_from_data(df)
          templates = load_grade_templates()
          signature = get_file_signature(df)
          saved_template = templates.get(signature, {})

          all_points_items = []
          score_sources = {}

          for col in df.columns:
              header = clean_header(col)
              if is_points_column(header):
                  question_text = find_related_question(header, list(df.columns))
                  saved_value = saved_template.get("max_scores", {}).get(header)
                  visible_score = extract_score_from_text(question_text)
                  detected_value = detected_scores.get(header)

                  if saved_value is not None and float(saved_value) > 0:
                      default_value = saved_value
                      source = "قالب محفوظ"
                  elif visible_score is not None:
                      default_value = visible_score
                      source = "درجة ظاهرة في السؤال"
                  elif detected_value is not None and float(detected_value) > 0:
                      default_value = detected_value
                      source = "درجة تلقائية من عمود Points"
                  else:
                      default_value = 0
                      source = "مدخل من المستخدم"

                  score_sources[header] = source

                  all_points_items.append({
                      "عمود الدرجة": header,
                      "السؤال": question_text,
                      "الدرجة الكبرى": default_value,
                      "المصدر": source,
                      "أعلى درجة موجودة في الملف": detected_value if detected_value is not None else "غير محدد",
                  })

          if not all_points_items:
              st.error("ما لقيت أعمدة Points في الملف. تأكدي أن ملف الاستجابات يحتوي أعمدة درجات.")
              max_scores = {}
              can_split = False
          else:
              st.info("البرنامج يضع الدرجات التلقائية الموجودة في أعمدة Points داخل خانة الدرجة الكبرى، وأي سؤال لا توجد له درجة تلقائية يبقى 0.00 لتدخلين درجته من الإجابة النموذجية.")
              max_scores = {}
              excluded_columns = []
              zero_confirmed = []

              for idx, item in enumerate(all_points_items, start=1):
                  current_value = item["الدرجة الكبرى"]
                  default_score = float(current_value) if current_value is not None else 0.0

                  with st.container(border=True):
                      st.markdown(f"**{idx}. عمود الدرجة:** `{item['عمود الدرجة']}`")
                      st.caption(f"السؤال المرتبط: {item['السؤال']}")
                      st.caption(f"مصدر الدرجة الحالية: {item['المصدر']}")

                      if default_score == 0:
                          confirm_zero = st.checkbox(
                              "تأكيد أن هذا العمود لا تُكتب له درجة كبرى / لا يُحسب",
                              value=False,
                              key=f"confirm_zero_{signature}_{idx}_{item['عمود الدرجة']}",
                          )
                      else:
                          confirm_zero = False

                      score = st.number_input(
                          "الدرجة الكبرى",
                          min_value=0.0,
                          value=default_score,
                          step=0.5,
                          key=f"score_{signature}_{idx}_{item['عمود الدرجة']}",
                      )

                      if score > 0:
                          final_score = int(score) if float(score).is_integer() else score
                          max_scores[item["عمود الدرجة"]] = final_score

                          if item["المصدر"] == "مدخل من المستخدم":
                              score_sources[item["عمود الدرجة"]] = "مدخل من المستخدم"
                          elif final_score != item["الدرجة الكبرى"]:
                              score_sources[item["عمود الدرجة"]] = "تم تعديله من المستخدم"
                          else:
                              score_sources[item["عمود الدرجة"]] = item["المصدر"]

                      elif confirm_zero:
                          zero_confirmed.append(item["عمود الدرجة"])
                          excluded_columns.append(item["عمود الدرجة"])
                          score_sources[item["عمود الدرجة"]] = "مؤكد يدويًا لا يُحسب / مقفل"

              missing_scores = [
                  item["عمود الدرجة"]
                  for item in all_points_items
                  if item["عمود الدرجة"] not in max_scores
                  and item["عمود الدرجة"] not in zero_confirmed
              ]

              can_split = len(missing_scores) == 0

              if st.button("💾 حفظ قالب الدرجات لهذا الملف"):
                  if can_split:
                      templates[signature] = {
                          "file_name": original_file_name,
                          "new_base_name": new_base_name,
                          "max_scores": max_scores,
                          "excluded_columns": excluded_columns,
                      }
                      save_grade_templates(templates)
                      st.success("تم حفظ قالب الدرجات بنجاح ✅")
                  else:
                      st.error("لازم تكتبين الدرجة الكبرى لكل أعمدة Points قبل الحفظ.")

              if not can_split:
                  st.warning("باقي أعمدة بدون درجة كبرى: " + "، ".join(missing_scores))

          if max_scores:
              total_exam_score = sum(float(v) for v in max_scores.values())
              if total_exam_score.is_integer():
                  total_exam_score = int(total_exam_score)
              st.success(f"📌 الدرجة النهائية للاختبار: {total_exam_score}")
              st.session_state["last_max_scores"] = max_scores
              st.session_state["last_exam_total_score"] = total_exam_score

          with st.expander("عرض الدرجات الكبرى المعتمدة"):
              if max_scores:
                  total_exam_score = sum(float(v) for v in max_scores.values())
                  if total_exam_score.is_integer():
                      total_exam_score = int(total_exam_score)

                  scores_df = pd.DataFrame([
                      {
                          "عمود الدرجة": k,
                          "الدرجة المعتمدة": v,
                          "المصدر": score_sources.get(k, "مدخل من المستخدم"),
                      }
                      for k, v in max_scores.items()
                  ])

                  if "excluded_columns" in locals() and excluded_columns:
                      excluded_df = pd.DataFrame([
                          {
                              "عمود الدرجة": col,
                              "الدرجة المعتمدة": 0,
                              "المصدر": score_sources.get(col, "مؤكد يدويًا لا يُحسب / مقفل"),
                          }
                          for col in excluded_columns
                      ])
                      scores_df = pd.concat([scores_df, excluded_df], ignore_index=True)

                  total_row = pd.DataFrame([
                      {
                          "عمود الدرجة": "المجموع الكلي / الدرجة النهائية للاختبار",
                          "الدرجة المعتمدة": total_exam_score,
                          "المصدر": "مجموع الدرجات الظاهرة والمدخلة",
                      }
                  ])

                  scores_df = pd.concat([scores_df, total_row], ignore_index=True)

                  st.dataframe(scores_df, use_container_width=True)
                  st.success(f"📌 الدرجة النهائية للاختبار: {total_exam_score}")
              else:
                  st.info("لا توجد درجات محددة حتى الآن.")

          st.markdown("---")
          st.info("سيتم إنشاء ورقة باسم (معلومات) داخل كل ملف، وفيها خانات فارغة لاسم المصححة واسم المدققة ليتم تعبئتها يدويًا داخل Excel.")
          corrector_name = ""
          auditor_name = ""

          if st.button("✂️ تقسيم الاستجابات وتنزيل الملفات", disabled=not can_split):
              zip_buffer = BytesIO()
              part_names = []
              responses_per_part_list = []

              with ZipFile(zip_buffer, "w") as zip_file:
                  for i in range(0, len(df), chunk_size):
                      chunk = df.iloc[i:i + chunk_size].copy()
                      chunk.insert(0, "رقم", range(1, len(chunk) + 1))

                      excel_buffer = BytesIO()
                      with pd.ExcelWriter(excel_buffer, engine="openpyxl") as writer:
                          chunk.to_excel(writer, index=False, sheet_name="Responses")

                      formatted_file = format_excel_file(
                          excel_buffer,
                          lock_sheet=True,
                          merge_mode=False,
                          extra_hidden_columns=extra_hidden_columns,
                          max_scores=max_scores,
                          excluded_columns=excluded_columns,
                          corrector_name=corrector_name,
                          auditor_name=auditor_name,
                      )

                      file_number = (i // chunk_size) + 1
                      part_file_name = f"{new_base_name}-{file_number}.xlsx"
                      part_names.append(part_file_name)
                      responses_per_part_list.append(str(len(chunk)))

                      zip_file.writestr(part_file_name, formatted_file.getvalue())

              zip_buffer.seek(0)
              part_count = len(part_names)

              log_operation(
                  operation_type="تقسيم",
                  original_file=original_file_name,
                  new_file_name=new_base_name,
                  part_names=part_names,
                  part_count=part_count,
                  response_count=len(df),
                  responses_per_part="، ".join(responses_per_part_list),
                  split_done="نعم",
                  merge_done="لا",
                  status="تم",
              )

              st.success(f"تم تقسيم الملف إلى {part_count} ملف ✅")
              st.download_button(
                  label="⬇️ تنزيل الملفات المقسمة ZIP",
                  data=zip_buffer,
                  file_name=f"{new_base_name}-split_files.zip",
                  mime="application/zip",
              )


# =========================
# تبويب التجميع
# =========================
with tab2:
    st.markdown("<h3>📥 تجميع ملفات الاستجابات المقسمة</h3>", unsafe_allow_html=True)

    uploaded_files = st.file_uploader(
        "ارفعي ملفات Excel المقسمة",
        type=["xlsx"],
        accept_multiple_files=True,
        key="merge_files",
    )

    if uploaded_files:
        sorted_files = sorted(uploaded_files, key=lambda f: get_part_number(f.name))
        detected_base_name = detect_base_name_from_parts(sorted_files)

        merge_base_name = st.text_input(
            "اسم ملف التجميع النهائي",
            value=detected_base_name,
            help="مثال: تقن106 — الناتج سيكون تقن106-مجمعة.xlsx",
        )
        merge_base_name = safe_filename(merge_base_name)

        all_data = []
        for file in sorted_files:
            df_part = pd.read_excel(file, engine="openpyxl")
            all_data.append(df_part)

        combined = pd.concat(all_data, ignore_index=True)
        combined, recalculated_total_col, recalculated_score_cols = recompute_total_points_dataframe(combined)

        st.success(f"تم رفع {len(uploaded_files)} ملف، وعدد الاستجابات بعد التجميع: {len(combined)}")
        if recalculated_total_col:
            st.info(f"✅ تم حساب الدرجة النهائية تلقائيًا من {len(recalculated_score_cols)} أعمدة درجات، وسيظهر عمود Total Points في ملف الرصد.")
        else:
            st.warning("⚠️ لم يتم العثور على أعمدة Points لحساب الدرجة النهائية. تأكدي من رفع ملفات التصحيح المقسمة الصحيحة.")
        st.dataframe(combined.head(), use_container_width=True)

        if st.button("📥 تجميع وتنزيل الملف"):
            output_excel = BytesIO()
            with pd.ExcelWriter(output_excel, engine="openpyxl") as writer:
                combined.to_excel(writer, index=False, sheet_name="Merged")

            formatted_merged = format_excel_file(
                output_excel,
                lock_sheet=False,
                merge_mode=True,
                extra_hidden_columns=[],
                max_scores={},
            )

            merged_file_name = f"{merge_base_name}-مجمعة.xlsx"

            log_operation(
                operation_type="تجميع",
                original_file="، ".join([f.name for f in sorted_files]),
                new_file_name=merge_base_name,
                part_names=[f.name for f in sorted_files],
                part_count=len(sorted_files),
                response_count=len(combined),
                responses_per_part="",
                split_done="لا",
                merge_done="نعم",
                merged_file_name=merged_file_name,
                status="تم",
            )

            st.success("تم تجميع الملف بنجاح ✅")
            st.download_button(
                label="⬇️ تنزيل ملف التجميع النهائي",
                data=formatted_merged,
                file_name=merged_file_name,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )

            # ملف الرصد — إيميل، اسم عربي، اسم إنجليزي، مدرسة، Total Points فقط
            def _get_col(df, keys):
                """ابحث عن عمود بأي من الكلمات المفتاحية"""
                for col in df.columns:
                    h = clean_header(col).lower()
                    if any(k in h for k in keys):
                        return col
                return None

            email_col   = _get_col(combined, ["email"])
            nameen_col  = _get_col(combined, ["^name$"]) or _get_col(combined, ["name"])
            namear_col  = _get_col(combined, ["الاسم الرباعي","اسم الطالبة","الاسم الثلاثي","الاسم الكامل","الاسم والرقم","بيانات الطالبة"])
            school_col  = detect_school_column(combined)
            total_col   = recalculated_total_col or _get_col(combined, ["total points"])

            # لو الاسم مدمج مع الرقم استخرج الاسم فقط (قبل " - ")
            def _extract_name(val):
                if pd.isna(val): return ""
                s = str(val)
                return s.split(" - ")[0].strip() if " - " in s else s.strip()

            if email_col or namear_col or total_col:
                rdf = pd.DataFrame()
                rdf["الإيميل"]        = combined[email_col].fillna("")  if email_col  else ""
                rdf["الاسم الإنجليزي"]= combined[nameen_col].fillna("") if nameen_col else ""
                rdf["الاسم العربي"]   = combined[namear_col].apply(_extract_name) if namear_col else ""
                rdf["المدرسة"]        = combined[school_col].fillna("") if school_col else ""
                rdf["Total Points"]   = combined[total_col].fillna("")  if total_col  else ""

                # تنسيق ملف الرصد
                roster_buf = BytesIO()
                with pd.ExcelWriter(roster_buf, engine="openpyxl") as wr:
                    rdf.to_excel(wr, index=False, sheet_name="الرصد")
                roster_buf.seek(0)
                wb_r = load_workbook(roster_buf)
                ws_r = wb_r.active
                ws_r.sheet_view.rightToLeft = True

                hdr_fill  = PatternFill("solid", fgColor="15396B")
                alt_fill  = PatternFill("solid", fgColor="EAF6FF")
                tot_fill  = PatternFill("solid", fgColor="FFF2CC")
                thin_b    = Border(
                    left=Side(style="thin", color="AAAAAA"),
                    right=Side(style="thin", color="AAAAAA"),
                    top=Side(style="thin", color="AAAAAA"),
                    bottom=Side(style="thin", color="AAAAAA"),
                )
                center_al = Alignment(horizontal="center", vertical="center", wrap_text=True)

                for cell in ws_r[1]:
                    cell.fill      = hdr_fill
                    cell.font      = Font(bold=True, size=13, color="FFFFFF")
                    cell.alignment = center_al
                    cell.border    = thin_b
                ws_r.row_dimensions[1].height = 40

                for row_idx, row in enumerate(ws_r.iter_rows(min_row=2), start=2):
                    bg = alt_fill if row_idx % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
                    for cell in row:
                        h_val = clean_header(ws_r.cell(1, cell.column).value).lower()
                        cell.fill      = tot_fill if "total" in h_val or "points" in h_val else bg
                        cell.font      = Font(size=12, bold=("total" in h_val or "points" in h_val))
                        cell.alignment = center_al
                        cell.border    = thin_b
                    ws_r.row_dimensions[row_idx].height = 28

                col_widths = {"الإيميل": 32, "الاسم الإنجليزي": 30,
                              "الاسم العربي": 28, "المدرسة": 32, "Total Points": 14}
                for cell in ws_r[1]:
                    key = clean_header(cell.value)
                    ws_r.column_dimensions[cell.column_letter].width = col_widths.get(key, 18)

                ws_r.freeze_panes = "A2"

                roster_out = BytesIO()
                wb_r.save(roster_out)
                roster_out.seek(0)

                st.download_button(
                    label="📊 تنزيل ملف الرصد (إيميل + اسم + درجة)",
                    data=roster_out,
                    file_name=f"{merge_base_name}-رصد.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )


# =========================
# تبويب تقرير الأدمن
# =========================
if tab3 is not None:
    with tab3:
        st.markdown("<h3>📊 تقرير الأدمن</h3>", unsafe_allow_html=True)

        if LOCAL_REPORT_FILE.exists():
            report_df = pd.read_excel(LOCAL_REPORT_FILE)
            st.dataframe(report_df, use_container_width=True)

            report_buffer = BytesIO()
            report_df.to_excel(report_buffer, index=False)
            report_buffer.seek(0)

            st.download_button(
                label="⬇️ تنزيل التقرير المحلي Excel",
                data=report_buffer,
                file_name="reports.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        else:
            st.info("لا يوجد تقرير محلي حتى الآن.")

        


# =========================
# فوتر أسفل الصفحة
# =========================
st.markdown(f"""
<style>
.footer-pro {{
    margin-top: 80px;
    padding: 22px 32px;
    background: linear-gradient(135deg, #0d3060 0%, #1a4f8a 100%);
    border-radius: 14px;
    display: flex;
    justify-content: space-between;
    align-items: center;
    gap: 16px;
    direction: rtl;
    box-shadow: 0 4px 20px rgba(13,48,96,0.2);
}}
.footer-pro .fp-item {{
    text-align: center;
    flex: 1;
}}
.footer-pro .fp-role {{
    color: rgba(255,255,255,0.6);
    font-size: 11px;
    font-weight: 600;
    letter-spacing: 0.5px;
    text-transform: uppercase;
    margin-bottom: 4px;
}}
.footer-pro .fp-name {{
    color: #ffffff;
    font-size: 14px;
    font-weight: 700;
}}
.footer-pro .fp-divider {{
    width: 1px;
    height: 36px;
    background: rgba(255,255,255,0.2);
    flex-shrink: 0;
}}
.footer-pro .fp-gold {{
    color: #e8c060;
}}
@media (max-width: 700px) {{
    .footer-pro {{
        flex-direction: column;
        text-align: center;
    }}
    .footer-pro .fp-divider {{ display: none; }}
}}
</style>

<div class="footer-pro">
    <div class="fp-item">
        <div class="fp-role">تصميم وبرمجة</div>
        <div class="fp-name fp-gold">أ. عفاف حسين</div>
    </div>
    <div class="fp-divider"></div>
    <div class="fp-item">
        <div class="fp-role">إشراف</div>
        <div class="fp-name">أ. أمينة الصائغ</div>
    </div>
    <div class="fp-divider"></div>
    <div class="fp-item">
        <div class="fp-role">رئيسة المركز</div>
        <div class="fp-name fp-gold">أ. خلود يعقوب بدر</div>
    </div>
</div>

""", unsafe_allow_html=True)

